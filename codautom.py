import streamlit as st
import pandas as pd
from docxtpl import DocxTemplate
from datetime import datetime
import os
import pdfplumber
import re
import openpyxl
from openpyxl.utils import coordinate_to_tuple
import io
import zipfile

# --- FUNÇÃO DE DATA ---
def data_extenso(dt):
    meses = {1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril", 5: "maio", 6: "junho",
             7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"}
    return f"{dt.day} de {meses[dt.month]} de {dt.year}"

# --- FILTRO ATÔMICO CONTRA RUÍDOS DE PÁGINA E TEXTOS VAZIOS ---
def limpar_texto_bloco(txt):
    if not txt: return ""
    linhas = txt.split('\n')
    linhas_limpas = []
    for l in linhas:
        l_strip = l.strip()
        if re.search(r'(?i)Página \d+ de \d+', l_strip): continue
        if re.search(r'(?i)UNIVERSIDADE FEDERAL DE SANTA MARIA', l_strip): continue
        if re.search(r'(?i)PROJETO NA ÍNTEGRA', l_strip): continue
        if re.search(r'(?i)Consulte em http', l_strip): continue
        if re.search(r'\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}', l_strip): continue
        if re.search(r'[A-F0-9]{4}(?:\.[A-F0-9]{4}){7}', l_strip): continue
        linhas_limpas.append(l)

    txt_final = " ".join([l.strip() for l in linhas_limpas if l.strip()])
    if txt_final.strip() in ["", "-", ".", "Não se aplica"]:
        return ""
    return txt_final.strip()

# --- DETECÇÃO AUTOMÁTICA DO INSTRUMENTO JURÍDICO ---
def identificar_instrumento_juridico(texto):
    texto = texto or ""

    m_instr = re.search(r'Instrumento jurídico celebrado\s*:\s*(.*?)(?:\n|$)', texto, re.IGNORECASE)
    if m_instr:
        valor = m_instr.group(1).strip().lower()
        if "acordo de cooperação técnica" in valor or "cooperação técnica" in valor or "cooperacao tecnica" in valor:
            return "Acordo de Cooperação Técnica (ACT)"
        if "contrato global" in valor:
            return "Contrato Global (CG)"
        if "acordo de parceria" in valor or "parceria" in valor:
            return "Acordo de Parceria (AP)"

    texto_low = texto.lower()
    if "acordo de cooperação técnica" in texto_low or "cooperação técnica" in texto_low or "cooperacao tecnica" in texto_low:
        return "Acordo de Cooperação Técnica (ACT)"
    if "contrato global" in texto_low:
        return "Contrato Global (CG)"
    if "acordo de parceria" in texto_low or "parceria" in texto_low:
        return "Acordo de Parceria (AP)"

    return "Acordo de Cooperação Técnica (ACT)"

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Raichu Pro", layout="wide")
st.title("Raichu Pro ⚡")

# --- ESTILIZAÇÃO CUSTOMIZADA (CSS) ---
st.markdown(
    """
    <style>
    div[data-testid="stRadio"] > label {
        font-size: 20px !important;
        font-weight: bold !important;
    }
    div[role="radiogroup"] p {
        font-size: 18px !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# ==============================================================================
# PASSO 1: SELEÇÃO DO PROCESSO
# ==============================================================================
st.markdown("### 1️⃣ Passo 1: Seleção do Processo e Relatório")
arquivo_pdf = st.file_uploader("Insira o seu relatório do projeto", type=["pdf"])

fundacoes_dados = {
    "FATEC": {"fundacao": "FATEC - Fundação de Apoio à Tecnologia e Ciência", "sigla_fundacao": "FATEC"},
    "FUNDEP": {"fundacao": "FUNDEP - Fundação de Desenvolvimento da Pesquisa", "sigla_fundacao": "FUNDEP"},
    "FAURGS": {"fundacao": "FAURGS - Fundação de Apoio à Universidade Federal do Rio Grande do Sul", "sigla_fundacao": "FAURGS"},
    "FDMS": {"fundacao": "FDMS - Fundação Delfim Mendes Silveira", "sigla_fundacao": "FDMS"}
}

dados_extraidos = {
    "titulo": "", "numero": "", "empresa": "", "data_inicio_proj": "", "data_termino_proj": "",
    "resumo": "", "objetivos": "", "justificativa_proj": "", "resultados": "", "importancia_projeto": "",
    "plano_gestao": "", "objetivo_estrategico": "", "inovacao_bool": "", "inovacao_potencial": "",
    "instrumento_juridico_pdf": "",
    "classificacoes_raw": [], "equipe_raw": [], "unidades_raw": [], "regioes_raw": [],
    "fundacao_sugerida": "FATEC", "tipo_processo_sugerido": "Acordo de Cooperação Técnica (ACT)"
}

texto_limpo = ""

# ==============================================================================
# MOTOR DE LEITURA DO PDF
# ==============================================================================
if arquivo_pdf:
    try:
        texto_completo = ""
        with pdfplumber.open(arquivo_pdf) as pdf:
            for page in pdf.pages:
                extraido = page.extract_text()
                if extraido: texto_completo += extraido + "\n"

        texto_limpo = re.sub(r'---\s*PAGE\s*\d+\s*---', '\n', texto_completo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'\d{2}/\d{2}/\d{4}\s\d{2}:\d{2}', '', texto_limpo)
        texto_limpo = re.sub(r'[A-F0-9]{4}(?:\.[A-F0-9]{4}){7}', '', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'Consulte em http[^\n]+', '', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'Registrado em:\s*\d{2}/\d{2}/\d{4}', '', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'UNIVERSIDADE FEDERAL DE SANTA MARIA - UFSM', '', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'PROJETO NA ÍNTEGRA', '', texto_limpo, flags=re.IGNORECASE)

        lixos_para_apagar = [
            r'(?i)PARTICIPANTE\s+V[ÍI]NCULO\s+CURSO/LOTA[ÇC][ÃA]O\s+FUN[ÇC][ÃA]O\s*\$\$\$',
            r'(?i)PARTICIPANTE\s+V[ÍI]NCULO\s+CURSO/LOTA[ÇC][ÃA]O\s+FUN[ÇC][ÃA]O',
            r'(?i)CH\s+DENTRO\s+CH\s+FORA\s+IN[ÍI]CIO\s+T[ÉE]RMINO\s+OBSERVA[ÇC][ÃA]O',
            r'(?i)UNIDADE\s+FUN[ÇC][ÃA]O\s+VALOR\s+IN[ÍI]CIO\s+T[ÉE]RMINO',
            r'(?i)TIPO\s+DE\s+CLASSIFICA[ÇC][ÃA]O\s+CLASSIFICA[ÇC][ÃA]O'
        ]
        for lixo in lixos_para_apagar:
            texto_limpo = re.sub(lixo, ' ', texto_limpo)

        for sigla in ["FATEC", "FUNDEP", "FAURGS", "FDMS"]:
            if re.search(r'\b' + sigla + r'\b', texto_limpo, re.IGNORECASE):
                dados_extraidos["fundacao_sugerida"] = sigla
                break

        dados_extraidos["tipo_processo_sugerido"] = identificar_instrumento_juridico(texto_limpo)

        def extrair(regex, group=1):
            m = re.search(regex, texto_limpo, re.IGNORECASE)
            return m.group(group).strip() if m else ""

        dados_extraidos["titulo"] = extrair(r'Título:\s*(.*?)\n')
        dados_extraidos["numero"] = extrair(r'Número:\s*(\d+)')
        dados_extraidos["data_inicio_proj"] = extrair(r'Início:\s*(\d{2}/\d{2}/\d{4})')
        dados_extraidos["data_termino_proj"] = extrair(r'Término:\s*(\d{2}/\d{2}/\d{4})')
        dados_extraidos["classificacao"] = extrair(r'Classificação:\s*(.*?)\n')
        dados_extraidos["empresa"] = extrair(r'(?:Financiador[a]?|Empresa|Cooperante|Financiador|Instituição):\s*(.*?)\n')
        dados_extraidos["instrumento_juridico_pdf"] = extrair(r'Instrumento jurídico celebrado:\s*([^\n]+)')

        m_coord = re.search(r'Responsável pelo projeto:\s*(.*?)\s*\(\s*(\d+)\s*\)', texto_limpo, re.IGNORECASE)
        if m_coord: dados_extraidos["coord_geral_pdf"] = {"nome": m_coord.group(1).strip(), "siape": m_coord.group(2).strip()}

        m_fisc = re.search(r'Fiscal:\s*(\d+)\s*-\s*(.*?)\s*\(', texto_limpo, re.IGNORECASE)
        if m_fisc: dados_extraidos["fiscal_pdf"] = {"siape": m_fisc.group(1).strip(), "nome": m_fisc.group(2).strip()}

        def extrair_bloco(inicio_regex, fins_regex):
            m_inicio = re.search(inicio_regex, texto_limpo, re.IGNORECASE)
            if not m_inicio: return ""
            idx = m_inicio.end()
            end_idx = len(texto_limpo)
            for f in fins_regex:
                mf = re.search(f, texto_limpo[idx:], re.IGNORECASE)
                if mf:
                    pos = idx + mf.start()
                    if pos < end_idx: end_idx = pos
            return texto_limpo[idx:end_idx].strip()

        dados_extraidos["resumo"] = limpar_texto_bloco(extrair_bloco(r'Resumo:', [r'Objetivos:']))
        dados_extraidos["objetivos"] = limpar_texto_bloco(extrair_bloco(r'Objetivos:', [r'Justificativa:']))
        dados_extraidos["justificativa_proj"] = limpar_texto_bloco(extrair_bloco(r'Justificativa:', [r'Resultados esperados:']))
        dados_extraidos["resultados"] = limpar_texto_bloco(extrair_bloco(r'Resultados esperados:', [r'PARTICIPANTES', r'PLANO DE GESTÃO', r'UNIDADES VINCULADAS']))

        cabecalho_combinado = r'PLANO DE GESTÃO\s*(?:-?\s*)?OBJETIVO ESTRATÉGICO'
        gestao_combined_raw = extrair_bloco(cabecalho_combinado, [r'\n\s*INOVAÇÃO', r'PROJETO POSSUI POTENCIAL', r'PROJETO POSSUI'])

        match_separa = re.match(r'(?i)^\s*(PDI\s*\d{4}-\d{4}\s*-\s*\w+)\s+(.*)', gestao_combined_raw)
        if match_separa:
            dados_extraidos["plano_gestao"] = limpar_texto_bloco(match_separa.group(1))
            dados_extraidos["objetivo_estrategico"] = limpar_texto_bloco(match_separa.group(2))
        else:
            if " - " in gestao_combined_raw:
                partes_gen = gestao_combined_raw.split(" - ", 1)
                dados_extraidos["plano_gestao"] = limpar_texto_bloco(partes_gen[0])
                dados_extraidos["objetivo_estrategico"] = limpar_texto_bloco(partes_gen[1])
            else:
                dados_extraidos["plano_gestao"] = limpar_texto_bloco(gestao_combined_raw)
                dados_extraidos["objetivo_estrategico"] = ""

        inov_bool = extrair_bloco(r'PROJETO POSSUI POTENCIAL DE INOVAÇÃO[^\n]*', [r'POTENCIAL DE INOVAÇÃO DO PROJETO'])
        if "Sim" in inov_bool: dados_extraidos["inovacao_bool"] = "Sim"
        elif "Não" in inov_bool or "Nao" in inov_bool: dados_extraidos["inovacao_bool"] = "Não"

        pot = extrair_bloco(r'POTENCIAL DE INOVAÇÃO DO PROJETO', [r'REGIÕES DE ATUAÇÃO', r'UNIDADES VINCULADAS', r'5 - UNIDADES'])
        dados_extraidos["inovacao_potencial"] = limpar_texto_bloco(pot)

        classif_blk = extrair_bloco(r'CLASSIFICAÇÕES', [r'PLANO DE GESTÃO'])
        for line in classif_blk.split('\n'):
            line = line.strip()
            if not line or "TIPO DE CLASSIFICAÇÃO" in line.upper() or line.upper() == "CLASSIFICAÇÃO" or "CLASSIFICAÇÕES" in line.upper(): continue

            m = re.search(r'\s+(\d{1,5}\.\d.*|\d{1,5}\s+-.*)', line)
            if m:
                tipo = line[:m.start()].strip()
                valor = line[m.start():].strip()
                dados_extraidos["classificacoes_raw"].append({"Tipo de Classificação": tipo, "Classificação": valor})
            else:
                if len(line) > 5:
                    dados_extraidos["classificacoes_raw"].append({"Tipo de Classificação": line, "Classificação": ""})

        bloco_participantes = extrair_bloco(r'PARTICIPANTES', [r'UNIDADES VINCULADAS\s*\n', r'CLASSIFICAÇÕES', r'REGIÕES DE ATUAÇÃO'])
        if not bloco_participantes:
            bloco_participantes = texto_limpo

        matches_participantes = list(re.finditer(r'(\d{5,15})\s*-\s*([A-ZÀ-Ÿ\s\']+?)\s*(?=[A-ZÀ-Ÿ][a-zà-ÿ]|UNIDADES VINCULADAS|CLASSIFICAÇÕES|$)', bloco_participantes))
        
        for i, match in enumerate(matches_participantes):
            siape = match.group(1).strip()
            nome_bruto = match.group(2).strip()

            nome_limpo = re.sub(r'\s+', ' ', nome_bruto).strip()
            corte_idx = len(nome_limpo)

            palavras_corte = [
                "VÍNCULO", "VINCULO", "CURSO", "LOTAÇÃO", "LOTACAO",
                "FUNÇÃO", "FUNCAO", "UNIDADE", "CLASSIFICA", "PARTICIPANTE",
                "CH DENTRO", "CH FORA", "INÍCIO", "INICIO", "TÉRMINO", "TERMINO",
                "DEPARTAMENTO", "OBSERVA", "VALOR", "TIPO", "$$$"
            ]

            for p in palavras_corte:
                idx = nome_limpo.upper().find(p)
                if idx != -1 and idx < corte_idx:
                    corte_idx = idx

            nome = nome_limpo[:corte_idx].strip(" -/")
            
            if not nome or len(nome) < 2:
                continue

            pos_inicio = match.end()
            pos_fim = matches_participantes[i+1].start() if i + 1 < len(matches_participantes) else pos_inicio + 400
            
            janela_texto = bloco_participantes[pos_inicio:pos_fim]

            janela_limpa = " ".join(janela_texto.split())
            vinculo, lotacao, funcao, bolsa, ch_d, ch_f, data_ini, data_fim = "Outro", "", "Participante", "Não", "0", "0", "", ""

            m_info = re.search(r'(Coordenador Administrativo|Coordenador|Estagiário|Colaborador|Fiscal|Participante|Membro|Pesquisador|Responsável Técnico|Responsável|Técnico|Bolsista)\s+(Sim|Não|Nao)[\s\S]*?(\d+)\s+(\d+)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})', janela_limpa, re.IGNORECASE)

            prefixos_ufsm_regex = [
                r"Estudante de Pós-\s*graduação", r"Estudante de Pós-Graduação", r"Estudante de Graduação", r"Estudante de graduação",
                r"Estudante de Ensino Médio", r"Técnico[- ]Administrativo em Educação", r"Técnico[- ]Administrativo", r"Tecnico Administrativo",
                r"Docente", r"Pesquisador", r"Participante Externo", r"Visitante", r"Estudante", r"Servidor", r"Outro"
            ]

            if m_info:
                text_before = janela_limpa[:m_info.start()].strip()
                text_after = janela_limpa[m_info.end():].strip()

                partes_nome_perdidas = []
                for w in text_after.split():
                    if w.lower() in ["administrativo", "em", "educação", "educacao", "técnico", "tecnico", "responsável", "responsavel"]:
                        break
                    if w.isupper() and len(w) > 1 and w not in ["SIM", "NÃO", "NAO"]:
                        partes_nome_perdidas.append(w)

                if partes_nome_perdidas:
                    nome_final_teste = nome + " " + " ".join(partes_nome_perdidas)
                    corte_idx_2 = len(nome_final_teste)
                    for p in palavras_corte:
                        idx = nome_final_teste.upper().find(p)
                        if idx != -1 and idx < corte_idx_2:
                            corte_idx_2 = idx
                    nome = nome_final_teste[:corte_idx_2].strip(" -/")

                vinculo_encontrado = False
                for prefixo in prefixos_ufsm_regex:
                    match_vinculo = re.match(r'(?i)^' + prefixo, text_before)
                    if match_vinculo:
                        vinculo = match_vinculo.group(0).strip()
                        vinculo = re.sub(r'\s+', ' ', vinculo).replace('- ', '-') 
                        lotacao = text_before[match_vinculo.end():].strip()
                        vinculo_encontrado = True
                        break

                if not vinculo_encontrado:
                    partes = text_before.split(" ", 1)
                    vinculo = partes[0] if len(partes) > 0 else "Outro"
                    lotacao = partes[1] if len(partes) > 1 else ""

                if "técnico" in vinculo.lower() or "técnico" in janela_limpa.lower() or "administrativo" in janela_limpa.lower():
                    if "administrativo" in janela_limpa.lower() or "educação" in janela_limpa.lower() or "educacao" in janela_limpa.lower():
                        vinculo = "Técnico-Administrativo em Educação"

                palavras_extras_lotacao = []
                for w in text_after.split():
                    if w.isupper() and w not in partes_nome_perdidas and w not in ["SIM", "NÃO", "NAO", "FISCAL", "COORDENADOR", "PESQUISADOR", "TÉCNICO", "TECNICO", "BOLSISTA"]:
                        palavras_extras_lotacao.append(w)
                if palavras_extras_lotacao:
                    lotacao = lotacao + " " + " ".join(palavras_extras_lotacao)

                lotacao_limpa = re.sub(r'\s+', ' ', lotacao).strip()
                corte_lot = len(lotacao_limpa)

                palavras_corte_lotacao = [
                    "UNIDADE", "CLASSIFICA", "PARTICIPANTE", "FUNÇÃO", "FUNCAO",
                    "VALOR", "INÍCIO", "INICIO", "TÉRMINO", "TERMINO", "OBSERVA", "TIPO", "$$$"
                ]

                for p in palavras_corte_lotacao:
                    idx = lotacao_limpa.upper().find(p)
                    if idx != -1 and idx < corte_lot:
                        corte_lot = idx

                lotacao = lotacao_limpa[:corte_lot].strip(" -/")

                funcao = m_info.group(1).title()
                bolsa = m_info.group(2).title()
                ch_d = m_info.group(3)
                ch_f = m_info.group(4)
                data_ini = m_info.group(5)
                data_fim = m_info.group(6)
            else:
                for prefixo in prefixos_ufsm_regex:
                    match_vinculo = re.match(r'(?i)^' + prefixo, janela_limpa)
                    if match_vinculo:
                        vinculo = match_vinculo.group(0).strip()
                        vinculo = re.sub(r'\s+', ' ', vinculo).replace('- ', '-')
                        remainder = janela_limpa[match_vinculo.end():]
                        lotacao = re.split(r'(?i)(Coordenador|Participante|Pesquisador|Estagiário|Colaborador|Membro|Fiscal|Responsável|Técnico|Bolsista)', remainder)[0].strip()
                        break

            dados_extraidos["equipe_raw"].append({
                "Nome": nome, "SIAPE": siape, "Vínculo": vinculo.title(), "Lotação": lotacao,
                "Função": funcao, "Bolsa": bolsa, "CH_D": ch_d, "CH_F": ch_f, "Início": data_ini, "Término": data_fim,
                "Chefia Imediata": "", "SIAPE Chefia": ""
            })

        unidades_blk = extrair_bloco(r'UNIDADES VINCULADAS\s*\n', [r'CLASSIFICAÇÕES', r'REGIÕES DE ATUAÇÃO', r'PARTICIPANTES'])
        for line in unidades_blk.split('\n'):
            if "UNIDADE" in line or not line.strip(): continue
            m_u = re.search(r'(.+?)\s+(Responsável|Colaborador|Fiscal|Coordenador|Membro|Financiador|Participante)\s+(?:([0-9.,\-]+)\s+)?(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})', line, re.IGNORECASE)
            if m_u:
                valor_u = m_u.group(3).strip() if m_u.group(3) else ""
                if valor_u == "-": valor_u = ""
                dados_extraidos["unidades_raw"].append({
                    "Unidade": m_u.group(1).strip(), "Função": m_u.group(2).strip(),
                    "Valor": valor_u, "Início": m_u.group(4).strip(), "Término": m_u.group(5).strip()
                })

        regioes_blk = extrair_bloco(r'REGIÕES DE ATUAÇÃO\s*\n', [r'DECLARAÇÃO', r'APROVAÇÕES', r'UNIDADES VINCULADAS'])
        for line in regioes_blk.split('\n'):
            line_str = line.strip()
            if "CIDADE" in line_str or not line_str: continue
            tokens = line_str.split()
            if len(tokens) >= 5:
                if re.match(r'\d{2}/\d{2}/\d{4}', tokens[-1]) and re.match(r'\d{2}/\d{2}/\d{4}', tokens[-2]):
                    fim = tokens[-1]
                    ini = tokens[-2]
                    pais = tokens[-3]
                    rem = tokens[:-3]
                    rem_str = " ".join(rem)
                    state_match = re.search(r'(Rio Grande do Sul|Santa Catarina|Paraná|Parana|São Paulo|Sao Paulo|RS|SC|PR|SP)$', rem_str, re.IGNORECASE)
                    if state_match:
                        uf = state_match.group(1).strip()
                        cidade = rem_str[:state_match.start()].strip()
                    else:
                        uf = rem[-1]
                        cidade = " ".join(rem[:-1])
                    dados_extraidos["regioes_raw"].append({
                        "Cidade": cidade, "UF": uf, "País": pais, "Início": ini, "Término": fim
                    })

    except Exception as e:
        st.error(f"❌ Erro no processamento do PDF: {str(e)}")

# ==============================================================================
# INTERFACE STREAMLIT - CONTINUAÇÃO DOS PASSOS
# ==============================================================================
if arquivo_pdf:
    st.markdown("---")

    st.markdown("### 2️⃣ Passo 2: Validação da Fundação")

    tipo_sugerido = dados_extraidos.get("tipo_processo_sugerido", "Acordo de Cooperação Técnica (ACT)")
    opcoes_processo = ["Acordo de Parceria (AP)", "Contrato Global (CG)", "Acordo de Cooperação Técnica (ACT)"]
    try:
        indice_padrao = opcoes_processo.index(tipo_sugerido)
    except ValueError:
        indice_padrao = 2

    st.info(f"📌 Instrumento jurídico identificado automaticamente no relatório: **{tipo_sugerido}**.")

    tipo_processo = st.radio(
        "Selecione o Tipo de Processo (ou confirme a sugestão):",
        opcoes_processo,
        index=indice_padrao,
        horizontal=True,
        key="tipo_processo_radio"
    )

    if tipo_processo == "Acordo de Cooperação Técnica (ACT)":
        st.info("💡 Processos do tipo **ACT** não necessitam de Fundação de Apoio.")
        status_fund, fund_sigla, ctx_fundacao = "Não possui", "ACT", {}
    else:
        fund_sugerida = dados_extraidos.get("fundacao_sugerida", "FATEC")

        st.info(f"🤖 O robô identificou que este projeto parece estar vinculado à fundação: **{fund_sugerida}**.")
        fundacao_correta = st.radio("A fundação identificada acima está correta?", ["Sim", "Não"], index=0, horizontal=True)

        ctx_fundacao = {}
        if fundacao_correta == "Sim":
            fund_sigla = fund_sugerida
            status_fund = "Já definida"
            ctx_fundacao = fundacoes_dados[fund_sigla]
        else:
            fund_sigla = st.selectbox("Por favor, selecione a fundação correta abaixo:", list(fundacoes_dados.keys()))
            status_fund = "Já definida"
            ctx_fundacao = fundacoes_dados[fund_sigla]

    st.markdown("---")
    st.markdown("### 3️⃣ Passo 3: Conferência e Edição de Dados")

    with st.expander("📝 Detalhes do Projeto e Textos Longos", expanded=True):
        st.warning("⚠️ **ATENÇÃO:** O robô preencheu os dados automaticamente com base no PDF. Por favor, confira todos os campos abaixo. Se algum dado estiver incorreto ou faltando, você pode **corrigir ou preencher manualmente** nestas caixas antes de gerar os documentos.")
        c1, c2 = st.columns(2)
        with c1:
            tit_proj = st.text_input("Nome do Projeto (Título)", value=dados_extraidos.get("titulo", ""))
            n_proj = st.text_input("Número do Registro GAP", value=dados_extraidos.get("numero", ""))
            resumo = st.text_area("Resumo do Projeto", value=dados_extraidos.get("resumo", ""), height=120)
            objetivos = st.text_area("Objetivos do Projeto", value=dados_extraidos.get("objetivos", ""), height=120)
            justificativa = st.text_area("Justificativa do Projeto", value=dados_extraidos.get("justificativa_proj", ""), height=120)
            importancia = st.text_area("Importância do Projeto", value=dados_extraidos.get("importancia_projeto", ""), height=80)
            justificativa_fund = st.text_area("Justificativa para escolha da Fundação", placeholder="Digite o motivo da escolha da fundação...", height=80)
        with c2:
            diretor_unidade = st.text_input("Diretor da Unidade")
            siape_diretor = st.text_input("SIAPE do Diretor")
            st.text_input("Classificação", value=dados_extraidos.get("classificacao", ""), disabled=True)
            data_termino_edit = st.text_input("Data de Término", value=dados_extraidos.get("data_termino_proj", ""))
            
            instrumento_juridico_edit = st.text_input("Instrumento Jurídico (Excel)", value=dados_extraidos.get("instrumento_juridico_pdf", ""))
            
            resultados = st.text_area("Resultados Esperados", value=dados_extraidos.get("resultados", ""), height=120)
            metas = st.text_area("Metas do Projeto (Opcional)", placeholder="Digite as metas do projeto...", height=120)

    st.markdown("---")
    st.subheader("👨‍🏫 Coordenador e Fiscal do Projeto")
    col_c1, col_c2, col_f1, col_f2 = st.columns(4)
    c_g_n = col_c1.text_input("Coordenador", value=dados_extraidos.get("coord_geral_pdf", {}).get("nome", "") if dados_extraidos.get("coord_geral_pdf") else "")
    c_g_s = col_c2.text_input("SIAPE Coord.", value=dados_extraidos.get("coord_geral_pdf", {}).get("siape", "") if dados_extraidos.get("coord_geral_pdf") else "")
    f_nome = col_f1.text_input("Fiscal", value=dados_extraidos.get("fiscal_pdf", {}).get("nome", "") if dados_extraidos.get("fiscal_pdf") else "")
    f_siape = col_f2.text_input("SIAPE Fiscal", value=dados_extraidos.get("fiscal_pdf", {}).get("siape", "") if dados_extraidos.get("fiscal_pdf") else "")

    st.markdown("---")
    col_adm1, col_adm2 = st.columns(2)
    nome_coord_adm = col_adm1.text_input("Coordenador Administrativo")
    siape_coord_adm = col_adm2.text_input("SIAPE Coord. Adm.")

    st.markdown("---")
    st.subheader("🏢 Empresas / Parceiras")
    st.info("Digite manualmente na caixinha abaixo o nome da empresa, ou das empresas. Ex: a FAURGS, o Banco do Brasil")
    num_empresas = st.number_input("Quantas empresas/instituições parceiras participam deste projeto?", min_value=1, max_value=10, value=1)

    empresas_lista = []
    for i in range(num_empresas):
        val_nome = dados_extraidos.get("empresa", "") if i == 0 else ""
        nome_emp = st.text_input(f"Nome da Empresa {i+1}", value=val_nome, key=f"emp_nome_unica_{i}")

        if nome_emp.strip():
            empresas_lista.append({"nome": nome_emp.strip()})

    st.markdown("---")
    st.write("### 📊 Tabelas Estruturadas Consolidadas")

    t1 = st.tabs(["👥 Equipe"])[0]

    with t1:
        st.warning("⚠️ **AVISO IMPORTANTE:** Preencha as colunas **'Chefia Imediata'** e **'SIAPE Chefia'** para cada participante clicando duas vezes no espaço vazio. Isso é obrigatório para as declarações de Carga Horária.")

        equipe_final = dados_extraidos["equipe_raw"].copy()
        if f_nome and not any(e["SIAPE"] == f_siape for e in equipe_final):
            equipe_final.append({"Nome": f_nome, "SIAPE": f_siape, "Vínculo": "Docente", "Lotação": "DEPARTAMENTO", "Função": "Fiscal", "CH_D": "0", "CH_F": "0", "Bolsa": "Não", "Início": "", "Término": "", "Chefia Imediata": "", "SIAPE Chefia": ""})

        df_equipe = pd.DataFrame(equipe_final).fillna("")
        df_equipe_edit = st.data_editor(df_equipe, num_rows="dynamic", key="ed_equipe", use_container_width=True)
        equipe_final = df_equipe_edit.fillna("").to_dict(orient="records")

    st.markdown("---")
    st.markdown("### 4️⃣ Passo 4: Geração de Documentos")
    st.write("Ao clicar no botão abaixo, o sistema irá preencher todos os documentos na nuvem e preparar um arquivo .ZIP para você baixar.")

    nomes_empresas_validas = [str(e["nome"]).strip() for e in empresas_lista if str(e["nome"]).strip()]
    if len(nomes_empresas_validas) == 0:
        st.warning("⚠️ ALERTA: Você não preencheu o nome de nenhuma empresa/parceira no campo '🏢 Empresas / Parceiras'. A chave no Word ficará vazia!")

    if st.button("🚀 Processar Documentos"):
        with st.spinner("⏳ Processando e gerando os documentos... Por favor, aguarde!"):
            logs = []
            estudantes_ignorados_log = []

            # ==========================================================================
            # 🎯 LÓGICA DAS MÚLTIPLAS EMPRESAS (PARA O WORD)
            # ==========================================================================
            if len(nomes_empresas_validas) == 0:
                texto_empresas = ""
            elif len(nomes_empresas_validas) == 1:
                texto_empresas = f" e {nomes_empresas_validas[0]}"
            elif len(nomes_empresas_validas) == 2:
                texto_empresas = f", {nomes_empresas_validas[0]} e {nomes_empresas_validas[1]}"
            else:
                texto_empresas = ", " + ", ".join(nomes_empresas_validas[:-1]) + f" e {nomes_empresas_validas[-1]}"
            # ==========================================================================

            base_instr = "Acordo de Cooperação Técnica"
            if tipo_processo == "Acordo de Parceria (AP)":
                base_instr = "Acordo de Parceria"
            elif tipo_processo == "Contrato Global (CG)":
                base_instr = "Contrato"

            sufixo_classificacao = dados_extraidos.get("classificacao", "").strip()
            for c in dados_extraidos["classificacoes_raw"]:
                if "caracterização das ações de extensão" in str(c.get("Tipo de Classificação", "")).lower():
                    val = str(c.get("Classificação", ""))
                    m_suf = re.search(r'[\d\.]+\s*-\s*(.*)', val)
                    if m_suf:
                        sufixo_classificacao = m_suf.group(1).strip()
                    else:
                        sufixo_classificacao = val.strip()
                    break
            
            texto_instrumento_completo = f"{base_instr} com {sufixo_classificacao}" if sufixo_classificacao else base_instr

            if tipo_processo == "Contrato Global (CG)": pasta_alvo = f"Modelos/AG/{fund_sigla}" if status_fund == "Já definida" else "Modelos/AG/SEM"
            elif tipo_processo == "Acordo de Parceria (AP)": pasta_alvo = f"Modelos/AP/{fund_sigla}" if status_fund == "Já definida" else "Modelos/AP/SEM"
            else: pasta_alvo = "Modelos/ACT"

            ctx_global = {
                "data_atual": data_extenso(datetime.now()), "dataatual": data_extenso(datetime.now()),
                "nome_projeto": tit_proj, "nomeprojeto": tit_proj,
                "titulo_projeto": tit_proj, "tituloprojeto": tit_proj,
                "n_projeto": n_proj, "nprojeto": n_proj,
                "classificacao": dados_extraidos["classificacao"],
                "instrumento_completo": instrumento_juridico_edit,
                "texto_empresas": texto_empresas,
                "nome_coord": c_g_n, "nomecoord": c_g_n,
                "siape_coord": c_g_s, "siapecoord": c_g_s,
                "nome_fiscal": f_nome, "nomefiscal": f_nome,
                "fiscal": f_nome,
                "siape_fiscal": f_siape, "siapefiscal": f_siape,
                "nome_coord_adm": nome_coord_adm, "nomecoordadm": nome_coord_adm,
                "siape_adm": siape_coord_adm, "siapeadm": siape_coord_adm,
                "membros": equipe_final, "objetivos": objetivos, "metas": metas,
                "justificativa": justificativa, "resultados": resultados,
                "importancia_projeto": importancia, "importanciaprojeto": importancia,
                "justificativa_fund": justificativa_fund, "justificativafund": justificativa_fund,
                "diretor_unidade": diretor_unidade, "diretorunidade": diretor_unidade,
                "siape_diretor": siape_diretor, "siapediretor": siape_diretor
            }
            ctx_global.update(ctx_fundacao)

            if not os.path.exists(pasta_alvo):
                st.error(f"❌ A pasta de modelos não foi encontrada: {pasta_alvo}")
            else:
                nome_pasta_principal = f"Documentos_Gerados_{fund_sigla}_{n_proj}"
                nome_pasta_principal = re.sub(r'[\\/*?:"<>|]', "", nome_pasta_principal)

                arquivos_na_pasta = [f for f in os.listdir(pasta_alvo) if not f.startswith("~$")]
                keywords_individuais = ["ch_dentro", "ch_fora", "conflito", "participante", "membro"]

                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

                    for arquivo in arquivos_na_pasta:
                        if arquivo.endswith(".docx"):
                            caminho_arquivo = os.path.join(pasta_alvo, arquivo)
                            nome_minusculo = arquivo.lower()
                            is_individual = any(kw in nome_minusculo for kw in keywords_individuais)

                            if is_individual:
                                for membro in equipe_final:
                                    if not membro.get("Nome") or str(membro.get("Nome")).strip() == "": continue
                                    
                                    vinculo_membro = str(membro.get("Vínculo", "")).lower()
                                    funcao_membro = str(membro.get("Função", "")).lower()
                                    
                                    if "estudante" in vinculo_membro or "bolsista" in funcao_membro or "estagiário" in funcao_membro or "estagiario" in funcao_membro:
                                        if arquivo == arquivos_na_pasta[0] or len(estudantes_ignorados_log) < sum("estudante" in str(m.get("Vínculo", "")).lower() for m in equipe_final):
                                            if membro.get("Nome") not in estudantes_ignorados_log:
                                                estudantes_ignorados_log.append(str(membro.get("Nome")))
                                        continue 

                                    ch_d_val = str(membro.get("CH_D", "0")).strip()
                                    ch_f_val = str(membro.get("CH_F", "0")).strip()

                                    if "ch_dentro" in nome_minusculo and ch_d_val in ["0", "0.0", "0,0", "-", ""]:
                                        continue
                                        
                                    if "ch_fora" in nome_minusculo and ch_f_val in ["0", "0.0", "0,0", "-", ""]:
                                        continue

                                    nome_limpo = re.sub(r'[^\w]', '_', str(membro.get("Nome")))[:40].strip('_')
                                    nome_doc_sem_ext = arquivo.replace(".docx", "")

                                    try:
                                        doc_ind = DocxTemplate(caminho_arquivo)
                                        ctx_membro = ctx_global.copy()
                                        ctx_membro.update(membro)
                                        ctx_membro["participante"] = membro.get("Nome", "")
                                        ctx_membro["siape"] = membro.get("SIAPE", "")
                                        ctx_membro["cargo"] = membro.get("Função", "")
                                        ctx_membro["ch_dentro"] = membro.get("CH_D", "0")
                                        ctx_membro["chdentro"] = membro.get("CH_D", "0")
                                        ctx_membro["ch_fora"] = membro.get("CH_F", "0")
                                        ctx_membro["chfora"] = membro.get("CH_F", "0")

                                        chefia_nome_val = str(membro.get("Chefia Imediata", ""))
                                        ctx_membro["chefia_imediata"] = chefia_nome_val
                                        ctx_membro["nome_chefia"] = chefia_nome_val
                                        ctx_membro["nomechefia"] = chefia_nome_val
                                        ctx_membro["chefia"] = chefia_nome_val
                                        ctx_membro["chefiaimediata"] = chefia_nome_val
                                        ctx_membro["nome_chefia_imediata"] = chefia_nome_val
                                        ctx_membro["nomechefiaimediata"] = chefia_nome_val

                                        ctx_membro["siape_chefia"] = str(membro.get("SIAPE Chefia", ""))
                                        ctx_membro["siapechefia"] = str(membro.get("SIAPE Chefia", ""))
                                        ctx_membro["siape_chefia_imediata"] = str(membro.get("SIAPE Chefia", ""))

                                        doc_ind.render(ctx_membro)
                                        doc_buffer = io.BytesIO()
                                        doc_ind.save(doc_buffer)
                                        zip_file.writestr(f"02_Documentos_Individuais/{nome_limpo}/{nome_limpo}_{nome_doc_sem_ext}.docx", doc_buffer.getvalue())
                                    except Exception as e:
                                        logs.append(f"Erro em {arquivo} para {membro.get('Nome')}: {str(e)}")

                            else:
                                try:
                                    doc = DocxTemplate(caminho_arquivo)
                                    doc.render(ctx_global)
                                    doc_buffer = io.BytesIO()
                                    doc.save(doc_buffer)
                                    zip_file.writestr(f"01_Documentos_Gerais/{arquivo}", doc_buffer.getvalue())
                                except Exception as e:
                                    logs.append(f"Erro ao processar arquivo geral {arquivo}: {str(e)}")

                    if arq_excel := next((f for f in arquivos_na_pasta if f.endswith(".xlsx")), None):
                        try:
                            caminho_excel = os.path.join(pasta_alvo, arq_excel)
                            wb = openpyxl.load_workbook(caminho_excel)
                            ws = wb["Plano de Trabalho"] if "Plano de Trabalho" in wb.sheetnames else wb.worksheets[0]

                            def escrever_excel(celula, valor):
                                val_str = str(valor).strip() if valor is not None else ""
                                if val_str in ["", "-", "None", "Não se aplica"]: val_str = None
                                try:
                                    r_row, r_col = coordinate_to_tuple(celula)
                                    for merged_range in list(ws.merged_cells.ranges):
                                        min_col, min_row, max_col, max_row = merged_range.bounds
                                        if min_col <= r_col <= max_col and min_row <= r_row <= max_row:
                                            intervalo = str(merged_range)
                                            ws.unmerge_cells(intervalo)
                                            ws.cell(row=min_row, column=min_col).value = val_str
                                            ws.merge_cells(intervalo)
                                            return
                                    ws.cell(row=r_row, column=r_col).value = val_str
                                except Exception as err:
                                    logs.append(f"Aviso na célula {celula}: {str(err)}")

                            nome_fiscal_excel = f_nome if str(f_nome).strip() != "" else "(Não possui)"
                            nome_coord_adm_excel = nome_coord_adm if str(nome_coord_adm).strip() != "" else "(Não possui)"

                            if tipo_processo == "Acordo de Cooperação Técnica (ACT)":
                                escrever_excel("C17", tit_proj)
                                escrever_excel("C19", data_termino_edit)
                                escrever_excel("C20", c_g_n)
                                escrever_excel("C21", c_g_s)
                                escrever_excel("C22", nome_fiscal_excel)
                                escrever_excel("C23", f_siape)
                                escrever_excel("C24", nome_coord_adm_excel)
                                escrever_excel("C25", siape_coord_adm)
                                escrever_excel("C26", n_proj)
                                escrever_excel("C27", dados_extraidos.get("classificacao", ""))
                                escrever_excel("C28", instrumento_juridico_edit)

                                escrever_excel("A32", resumo)
                                escrever_excel("A36", objetivos)
                                escrever_excel("A40", justificativa)
                                escrever_excel("A44", resultados)

                            else:
                                escrever_excel("C25", tit_proj)
                                escrever_excel("C27", data_termino_edit)
                                escrever_excel("C28", c_g_n)
                                escrever_excel("C29", c_g_s)
                                escrever_excel("C30", nome_fiscal_excel)
                                escrever_excel("C31", f_siape)
                                escrever_excel("C32", nome_coord_adm_excel)
                                escrever_excel("C33", siape_coord_adm)
                                escrever_excel("C34", n_proj)
                                escrever_excel("C35", dados_extraidos.get("classificacao", ""))
                                escrever_excel("C36", instrumento_juridico_edit)
                                
                                escrever_excel("A40", resumo)
                                escrever_excel("A44", objetivos)
                                escrever_excel("A48", justificativa)
                                escrever_excel("A52", resultados)

                                equipe_excel = [p for p in equipe_final if p.get("Função", "") != "Fiscal" and str(p.get("Nome", "")).strip() != ""]
                                for idx, p in enumerate(equipe_excel):
                                    linha = 104 + idx
                                    if linha > 145:
                                        break
                                    escrever_excel(f"C{linha}", p.get("Nome", ""))
                                    escrever_excel(f"F{linha}", p.get("SIAPE", ""))

                            excel_buffer = io.BytesIO()
                            wb.save(excel_buffer)
                            zip_file.writestr(f"01_Documentos_Gerais/{arq_excel}", excel_buffer.getvalue())
                        except Exception as e:
                            logs.append(f"❌ Erro crítico no Excel Mestre: {str(e)}")

                if logs:
                    st.warning("⚠️ Foram gerados arquivos, mas ocorreram alguns avisos:")
                    for l in logs: st.error(l)
                else:
                    st.success("🔥 Documentos gerados e empacotados com Sucesso Absoluto!")
                    
                    if estudantes_ignorados_log:
                        st.info(f"🎓 **Filtro Automático:** O sistema bloqueou propositalmente a geração de documentos individuais (Carga Horária) para **{len(estudantes_ignorados_log)} estudante(s)/bolsista(s)**: {', '.join(estudantes_ignorados_log)}.")
                    
                    st.warning("📝 **LEMBRETE:** Após baixar e descompactar o ZIP, todos os documentos estarão em **Word (.docx)** e **Excel (.xlsx)**. Pode abri-los e editar qualquer texto normalmente no seu computador.")

                st.session_state['zip_data'] = zip_buffer.getvalue()
                st.session_state['zip_name'] = f"{nome_pasta_principal}.zip"

    if 'zip_data' in st.session_state:
        st.download_button(
            label="⬇️ CLIQUE AQUI PARA BAIXAR OS DOCUMENTOS (.ZIP)",
            data=st.session_state['zip_data'],
            file_name=st.session_state['zip_name'],
            mime="application/zip",
            type="primary"
        )

st.markdown("<br><hr>", unsafe_allow_html=True)
st.markdown("<div style='text-align: center; color: #888888; padding: 10px; font-size: 14px;'>⚡ <b>Raichu Pro</b> | Desenvolvido por Julio Maia 👨‍💻</div>", unsafe_allow_html=True)
