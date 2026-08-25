import os
import io
import re
import zipfile
import openpyxl
from openpyxl.utils import coordinate_to_tuple
from openpyxl.styles import Alignment
from fastapi import FastAPI, HTTPException, Response
from pydantic import BaseModel
from docxtpl import DocxTemplate
from datetime import datetime

# ==========================================
# INICIALIZAÇÃO
# ==========================================
app = FastAPI()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ==========================================
# MODELOS PYDANTIC (A estrutura robusta de dados)
# ==========================================
class DadosProjeto(BaseModel):
    titulo: str = ""
    numero: str = ""
    resumo: str = ""
    objetivos: str = ""
    justificativa: str = ""
    importancia: str = ""
    justificativa_fund: str = ""
    resultados: str = ""
    metas: str = ""
    classificacao: str = ""
    instrumento_juridico: str = ""
    data_termino: str = ""
    tipo_processo: str = ""
    fundacao_sigla: str = ""
    status_fund: str = ""
    classificacoes_raw: list = []

class Pessoas(BaseModel):
    coordenador: str = ""
    siape_coord: str = ""
    fiscal: str = ""
    siape_fiscal: str = ""
    coord_adm: str = ""
    siape_adm: str = ""
    diretor: str = ""
    siape_diretor: str = ""

class PayloadCompleto(BaseModel):
    dados_projeto: DadosProjeto
    pessoas: Pessoas
    empresas: list = []
    equipe: list = []

def data_extenso(dt):
    meses = {1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril", 5: "maio", 6: "junho",
             7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"}
    return f"{dt.day} de {meses[dt.month]} de {dt.year}"

# ==========================================
# ROTA MESTRE DE GERAÇÃO
# ==========================================
@app.post("/gerar-zip-completo/")
async def gerar_zip_completo(payload: PayloadCompleto):
    try:
        d_proj = payload.dados_projeto
        pess = payload.pessoas
        equipe_final = payload.equipe
        nomes_empresas_validas = payload.empresas
        fund_sigla = d_proj.fundacao_sigla

        if len(nomes_empresas_validas) == 0:
            texto_empresas = ""
        elif len(nomes_empresas_validas) == 1:
            texto_empresas = f" e {nomes_empresas_validas[0]}"
        elif len(nomes_empresas_validas) == 2:
            texto_empresas = f", {nomes_empresas_validas[0]} e {nomes_empresas_validas[1]}"
        else:
            texto_empresas = ", " + ", ".join(nomes_empresas_validas[:-1]) + f" e {nomes_empresas_validas[-1]}"

        base_instr = "Acordo de Cooperação Técnica"
        if d_proj.tipo_processo == "Acordo de Parceria (AP)": base_instr = "Acordo de Parceria"
        elif d_proj.tipo_processo == "Contrato Global (CG)": base_instr = "Contrato"

        sufixo_classificacao = d_proj.classificacao.strip()
        for c in d_proj.classificacoes_raw:
            if isinstance(c, dict) and "caracterização das ações de extensão" in str(c.get("Tipo de Classificação", "")).lower():
                val = str(c.get("Classificação", ""))
                m_suf = re.search(r'[\d\.]+\s*-\s*(.*)', val)
                sufixo_classificacao = m_suf.group(1).strip() if m_suf else val.strip()
                break
        
        texto_instrumento_completo = f"{base_instr} com {sufixo_classificacao}" if sufixo_classificacao else base_instr

        if d_proj.tipo_processo == "Contrato Global (CG)": 
            pasta_alvo = os.path.join(BASE_DIR, "Modelos", "AG", fund_sigla) if d_proj.status_fund == "Já definida" else os.path.join(BASE_DIR, "Modelos", "AG", "SEM")
        elif d_proj.tipo_processo == "Acordo de Parceria (AP)": 
            pasta_alvo = os.path.join(BASE_DIR, "Modelos", "AP", fund_sigla) if d_proj.status_fund == "Já definida" else os.path.join(BASE_DIR, "Modelos", "AP", "SEM")
        else: 
            pasta_alvo = os.path.join(BASE_DIR, "Modelos", "ACT")

        if not os.path.exists(pasta_alvo):
            raise HTTPException(status_code=404, detail=f"Pasta de modelos não encontrada: {pasta_alvo}")

        # MAPEAMENTO TOTAL DE VARIÁVEIS DO WORD
        ctx_global = {
            "data_atual": data_extenso(datetime.now()), "dataatual": data_extenso(datetime.now()),
            "nome_projeto": d_proj.titulo, "nomeprojeto": d_proj.titulo,
            "titulo_projeto": d_proj.titulo, "tituloprojeto": d_proj.titulo,
            "titulo": d_proj.titulo,
            "n_projeto": d_proj.numero, "nprojeto": d_proj.numero,
            "numero": d_proj.numero,
            "classificacao": d_proj.classificacao,
            "instrumento_completo": d_proj.instrumento_juridico,
            "texto_empresas": texto_empresas,
            "nome_coord": pess.coordenador, "nomecoord": pess.coordenador,
            "siape_coord": pess.siape_coord, "siapecoord": pess.siape_coord,
            "nome_fiscal": pess.fiscal, "nomefiscal": pess.fiscal,
            "fiscal": pess.fiscal,
            "siape_fiscal": pess.siape_fiscal, "siapefiscal": pess.siape_fiscal,
            "nome_coord_adm": pess.coord_adm, "nomecoordadm": pess.coord_adm,
            "siape_adm": pess.siape_adm, "siapeadm": pess.siape_adm,
            "membros": equipe_final, 
            "objetivos": d_proj.objetivos, "metas": d_proj.metas,
            "justificativa": d_proj.justificativa, "resultados": d_proj.resultados,
            "importancia_projeto": d_proj.importancia, "importanciaprojeto": d_proj.importancia,
            "justificativa_fund": d_proj.justificativa_fund, "justificativafund": d_proj.justificativa_fund,
            "diretor_unidade": pess.diretor, "diretorunidade": pess.diretor,
            "siape_diretor": pess.siape_diretor, "siapediretor": pess.siape_diretor,
            "sigla_fundacao": fund_sigla
        }

        arquivos_na_pasta = [f for f in os.listdir(pasta_alvo) if not f.startswith("~$")]
        keywords_individuais = ["ch_dentro", "ch_fora", "conflito", "participante", "membro"]

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

            for arquivo in arquivos_na_pasta:
                caminho_arquivo = os.path.join(pasta_alvo, arquivo)
                
                if arquivo.endswith(".docx"):
                    nome_minusculo = arquivo.lower()
                    is_individual = any(kw in nome_minusculo for kw in keywords_individuais)

                    if is_individual:
                        for membro in equipe_final:
                            if not membro.get("Nome") or str(membro.get("Nome")).strip() == "": continue
                            
                            vinculo_membro = str(membro.get("Vínculo", "")).lower()
                            funcao_membro = str(membro.get("Função", "")).lower()
                            
                            if "estudante" in vinculo_membro or "bolsista" in funcao_membro or "estagiário" in funcao_membro or "estagiario" in funcao_membro:
                                continue 

                            ch_d_val = str(membro.get("CH_D", "0")).strip()
                            ch_f_val = str(membro.get("CH_F", "0")).strip()

                            if "ch_dentro" in nome_minusculo and ch_d_val in ["0", "0.0", "0,0", "-", ""]: continue
                            if "ch_fora" in nome_minusculo and ch_f_val in ["0", "0.0", "0,0", "-", ""]: continue

                            nome_limpo = re.sub(r'[^\w]', '_', str(membro.get("Nome")))[:40].strip('_')
                            nome_doc_sem_ext = arquivo.replace(".docx", "")

                            try:
                                doc_ind = DocxTemplate(caminho_arquivo)
                                ctx_membro = ctx_global.copy()
                                ctx_membro.update(membro)
                                
                                # Alias à prova de falhas para Carga Horária e Conflito
                                ctx_membro["participante"] = membro.get("Nome", "")
                                ctx_membro["nome"] = membro.get("Nome", "")
                                ctx_membro["Nome"] = membro.get("Nome", "")
                                ctx_membro["siape"] = membro.get("SIAPE", "")
                                ctx_membro["cargo"] = membro.get("Função", "")
                                
                                ctx_membro["ch_dentro"] = membro.get("CH_D", "0")
                                ctx_membro["chdentro"] = membro.get("CH_D", "0")
                                ctx_membro["ch_fora"] = membro.get("CH_F", "0")
                                ctx_membro["chfora"] = membro.get("CH_F", "0")

                                chefia_nome_val = str(membro.get("Chefia Imediata", ""))
                                ctx_membro["chefia_imediata"] = chefia_nome_val
                                ctx_membro["nome_chefia"] = chefia_nome_val
                                ctx_membro["nomechefia"] = chefia_nome_val
                                ctx_membro["chefia"] = chefia_nome_val
                                ctx_membro["chefiaimediata"] = chefia_nome_val

                                siape_chefia_val = str(membro.get("SIAPE Chefia", ""))
                                ctx_membro["siape_chefia"] = siape_chefia_val
                                ctx_membro["siapechefia"] = siape_chefia_val
                                ctx_membro["siape_chefia_imediata"] = siape_chefia_val

                                doc_ind.render(ctx_membro)
                                doc_buffer_ind = io.BytesIO()
                                doc_ind.save(doc_buffer_ind)
                                zip_file.writestr(f"02_Documentos_Individuais/{nome_limpo}/{nome_limpo}_{nome_doc_sem_ext}.docx", doc_buffer_ind.getvalue())
                            except Exception as e:
                                print(f"Aviso DOCX IND {arquivo}: {str(e)}")
                    else:
                        try:
                            doc = DocxTemplate(caminho_arquivo)
                            doc.render(ctx_global)
                            doc_buffer = io.BytesIO()
                            doc.save(doc_buffer)
                            zip_file.writestr(f"01_Documentos_Gerais/{arquivo}", doc_buffer.getvalue())
                        except Exception as e:
                            print(f"Aviso DOCX GERAL {arquivo}: {str(e)}")

                elif arquivo.endswith(".xlsx"):
                    try:
                        wb = openpyxl.load_workbook(caminho_arquivo)
                        ws = wb["Plano de Trabalho"] if "Plano de Trabalho" in wb.sheetnames else wb.worksheets[0]

                        def escrever_excel(celula, valor):
                            val_str = str(valor).strip() if valor is not None else ""
                            if val_str in ["", "-", "None", "Não se aplica"]: val_str = None
                            try:
                                r_row, r_col = coordinate_to_tuple(celula)
                                
                                if val_str and len(val_str) > 0:
                                    qtd_quebras = val_str.count('\n')
                                    linhas_estimadas = (len(val_str) / 110.0) + qtd_quebras
                                    if linhas_estimadas < 1: linhas_estimadas = 1
                                    altura_calculada = (linhas_estimadas * 15) + 10
                                    altura_atual = ws.row_dimensions[r_row].height
                                    if altura_atual is None or altura_calculada > altura_atual:
                                        ws.row_dimensions[r_row].height = altura_calculada

                                for merged_range in list(ws.merged_cells.ranges):
                                    min_col, min_row, max_col, max_row = merged_range.bounds
                                    if min_col <= r_col <= max_col and min_row <= r_row <= max_row:
                                        intervalo = str(merged_range)
                                        ws.unmerge_cells(intervalo)
                                        cel_alvo = ws.cell(row=min_row, column=min_col)
                                        cel_alvo.value = val_str
                                        cel_alvo.alignment = Alignment(wrap_text=True, vertical='top')
                                        ws.merge_cells(intervalo)
                                        return
                                        
                                cel_alvo = ws.cell(row=r_row, column=r_col)
                                cel_alvo.value = val_str
                                cel_alvo.alignment = Alignment(wrap_text=True, vertical='top')
                            except Exception:
                                pass

                        nome_fiscal_excel = pess.fiscal if str(pess.fiscal).strip() != "" else "(Não possui)"
                        nome_coord_adm_excel = pess.coord_adm if str(pess.coord_adm).strip() != "" else "(Não possui)"

                        if d_proj.tipo_processo == "Acordo de Cooperação Técnica (ACT)":
                            escrever_excel("C17", d_proj.titulo)
                            escrever_excel("C19", d_proj.data_termino)
                            escrever_excel("C20", pess.coordenador)
                            escrever_excel("C21", pess.siape_coord)
                            escrever_excel("C22", nome_fiscal_excel)
                            escrever_excel("C23", pess.siape_fiscal)
                            escrever_excel("C24", nome_coord_adm_excel)
                            escrever_excel("C25", pess.siape_adm)
                            escrever_excel("C26", d_proj.numero)
                            escrever_excel("C27", d_proj.classificacao)
                            escrever_excel("C28", d_proj.instrumento_juridico)
                            escrever_excel("A32", d_proj.resumo)
                            escrever_excel("A36", d_proj.objetivos)
                            escrever_excel("A40", d_proj.justificativa)
                            escrever_excel("A44", d_proj.resultados)
                        else:
                            escrever_excel("C28", d_proj.titulo)
                            escrever_excel("C33", pess.coordenador)
                            escrever_excel("C37", nome_fiscal_excel)
                            escrever_excel("C39", nome_coord_adm_excel)
                            escrever_excel("C41", d_proj.numero)
                            escrever_excel("C42", texto_instrumento_completo)
                            escrever_excel("A46", d_proj.resumo)
                            escrever_excel("A50", d_proj.objetivos)
                            escrever_excel("A54", d_proj.resultados)

                        excel_buffer = io.BytesIO()
                        wb.save(excel_buffer)
                        zip_file.writestr(f"01_Documentos_Gerais/{arquivo}", excel_buffer.getvalue())
                    except Exception as e:
                        print(f"Erro no Excel: {str(e)}")

            # === O SEGREDO ESTÁ NO RECUO DESSAS LINHAS ===
        
        # Como essas linhas estão fora do "with zipfile", o Python fecha e salva o ZIP primeiro!
        zip_buffer.seek(0)
        nome_zip_saida = f"Documentos_Gerados_{fund_sigla}_{d_proj.numero or 'Projeto'}.zip"

        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={nome_zip_saida}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro crítico: {str(e)}")